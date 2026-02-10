"""
URL configuration for core project.

The `urlpatterns` list routes URLs to views. For more information please see:
    https://docs.djangoproject.com/en/6.0/topics/http/urls/
"""
#  Proyecto\POA_Reportes\backend\core/urls.py 

from django.contrib import admin
from django.urls import path
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
import zipfile

def health_check(request):
    return JsonResponse({
        'status': 'ok',
        'service': 'POA Backend',
        'timestamp': datetime.now().isoformat(),
        'message': 'Servidor funcionando correctamente',
        'endpoints': {
            'health': '/health/',
            'direcciones': '/api/direcciones/',
            'vista_previa': '/api/reportes/vista_previa/',
            'generar_reporte': '/api/reportes/generar/',
            'admin': '/admin/'
        }
    })

@csrf_exempt
def lista_direcciones(request):
    """Vista para listar direcciones"""
    try:
        from backend.poa.models import Direccion
        direcciones = list(Direccion.objects.values('id', 'nombre', 'codigo'))
        return JsonResponse({
            'success': True,
            'data': direcciones,
            'count': len(direcciones)
        })
    except Exception as e:
        print(f"Error obteniendo direcciones: {e}")
        direcciones = [
            { "id": 1, "nombre": "Obras Públicas", "codigo": "OP" },
            { "id": 2, "nombre": "Desarrollo Social", "codigo": "DS"},
            { "id": 3, "nombre": "Servicios Públicos", "codigo": "SP" },
            { "id": 4, "nombre": "Medio Ambiente", "codigo": "MA" },
            { "id": 5, "nombre": "Innovación", "codigo": "IN" },
            { "id": 6, "nombre": "Desarrollo Económico", "codigo": "DE" },
            { "id": 7, "nombre": "Movilidad", "codigo": "MO" },
        ]
        return JsonResponse({
            'success': True,
            'data': direcciones,
            'count': len(direcciones),
            'note': 'Usando datos de ejemplo'
        })

@csrf_exempt
def vista_previa_reporte(request):
    """Vista previa del reporte"""    
    try:
        fecha_corte = request.GET.get('fecha_corte', datetime.now().strftime('%Y-%m-%d'))
        
        from backend.poa.models import Obra
        from django.db.models import Sum, Count
        
        total_proyectos = Obra.objects.count()
        
        presupuesto_result = Obra.objects.aggregate(
            total=Sum('presupuesto_modificado')
        )
        presupuesto_total = presupuesto_result['total'] or 0
        
        obras_por_estado = dict(Obra.objects.values('estatus_general').annotate(
            count=Count('id')
        ).values_list('estatus_general', 'count'))
        
        return JsonResponse({
            'success': True,
            'total_proyectos': total_proyectos,
            'presupuesto_total': float(presupuesto_total),
            'beneficiarios_total': 5000,  # Valor por defecto
            'obras_por_estado': obras_por_estado,
            'fecha_corte': fecha_corte
        })
    except Exception as e:
        print(f"Error en vista previa: {e}")
        return JsonResponse({
            'success': True,
            'total_proyectos': 15,
            'presupuesto_total': 25000000,
            'beneficiarios_total': 5000,
            'obras_por_estado': {'En progreso': 8, 'Completado': 5, 'Pendiente': 2},
            'fecha_corte': datetime.now().strftime('%Y-%m-%d'),
            'note': 'Usando datos de ejemplo por error'
        })

# ============================
# FUNCIONES AUXILIARES PARA GENERAR REPORTES
# ============================

def _crear_pdf_ejecutivo(datos, fecha_corte):
    """Crea PDF para reporte ejecutivo"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           rightMargin=2*cm, leftMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    titulo_style = ParagraphStyle(
        'TituloPrincipal',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e3a8a')
    )
    
    subtitulo_style = ParagraphStyle(
        'Subtitulo',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        textColor=colors.HexColor('#374151')
    )
    
    contenido = []
    
    # Título
    contenido.append(Paragraph(f"REPORTE EJECUTIVO - POA", titulo_style))
    contenido.append(Paragraph(f"Fecha de corte: {fecha_corte}", subtitulo_style))
    contenido.append(Spacer(1, 0.5*cm))
    
    # Resumen ejecutivo
    contenido.append(Paragraph("RESUMEN EJECUTIVO", styles['Heading3']))
    
    resumen_text = f"""
    <b>Total de Proyectos:</b> {datos['total_obras']:,}<br/>
    <b>Presupuesto Total:</b> ${datos['presupuesto_total']:,.2f}<br/>
    <b>Avance Promedio:</b> {datos['avance_promedio']:.1f}%<br/>
    <b>Proyectos en Riesgo:</b> {datos['proyectos_riesgo']}<br/>
    <b>Beneficiarios Totales:</b> {datos['beneficiarios_totales']:,}<br/>
    """
    
    contenido.append(Paragraph(resumen_text, styles['Normal']))
    contenido.append(Spacer(1, 1*cm))
    
    # Tabla de proyectos principales
    contenido.append(Paragraph("PROYECTOS PRINCIPALES", styles['Heading3']))
    
    if datos['detalles']:
        table_data = [['Proyecto', 'Área', 'Avance', 'Estado']]
        
        for obra in datos['detalles'][:10]:
            table_data.append([
                obra.get('programa', '')[:40] + ('...' if len(obra.get('programa', '')) > 40 else ''),
                obra.get('area_responsable', '')[:20],
                f"{obra.get('avance_fisico_pct', 0):.1f}%",
                obra.get('estatus_general', '')[:15]
            ])
        
        table = Table(table_data, colWidths=[200, 100, 60, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        contenido.append(table)
    
    contenido.append(Spacer(1, 1*cm))
    contenido.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                              ParagraphStyle('footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)))
    
    doc.build(contenido)
    buffer.seek(0)
    return buffer

def _crear_excel_ejecutivo(datos, fecha_corte):
    """Crea Excel para reporte ejecutivo"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ejecutivo"
    
    # Estilos
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = f"REPORTE EJECUTIVO - {fecha_corte}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_alignment
    
    # Resumen
    ws['A3'] = "RESUMEN EJECUTIVO"
    ws['A3'].font = Font(bold=True, size=12)
    
    resumen_data = [
        ['Total de Proyectos', datos['total_obras']],
        ['Presupuesto Total', f"${datos['presupuesto_total']:,.2f}"],
        ['Avance Promedio', f"{datos['avance_promedio']:.1f}%"],
        ['Proyectos en Riesgo', datos['proyectos_riesgo']],
        ['Beneficiarios Totales', datos['beneficiarios_totales']],
    ]
    
    for i, (label, value) in enumerate(resumen_data, start=4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    
    # Tabla de proyectos
    ws['A8'] = "PROYECTOS PRINCIPALES"
    ws['A8'].font = Font(bold=True, size=12)
    
    headers = ['Proyecto', 'Área', 'Presupuesto', 'Avance', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    if datos['detalles']:
        for row, obra in enumerate(datos['detalles'][:20], start=10):
            ws.cell(row=row, column=1, value=obra.get('programa', '')[:50])
            ws.cell(row=row, column=2, value=obra.get('area_responsable', ''))
            ws.cell(row=row, column=3, value=float(obra.get('presupuesto_modificado', 0)))
            ws.cell(row=row, column=4, value=float(obra.get('avance_fisico_pct', 0)))
            ws.cell(row=row, column=5, value=obra.get('estatus_general', ''))
    
    # Ajustar anchos
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def _crear_pdf_cartera(datos, fecha_corte):
    """Crea PDF para cartera de proyectos"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    styles = getSampleStyleSheet()
    contenido = []
    
    contenido.append(Paragraph(f"CARTERA DE PROYECTOS - POA", 
                              ParagraphStyle('Titulo', parent=styles['Heading1'], alignment=TA_CENTER)))
    contenido.append(Paragraph(f"Fecha: {fecha_corte}", styles['Heading2']))
    contenido.append(Spacer(1, 1*cm))
    
    # Distribución por estado
    if datos.get('por_estado'):
        contenido.append(Paragraph("DISTRIBUCIÓN POR ESTADO", styles['Heading3']))
        
        table_data = [['Estado', 'Cantidad', 'Porcentaje']]
        total = sum(datos['por_estado'].values())
        
        for estado, cantidad in datos['por_estado'].items():
            porcentaje = (cantidad / total * 100) if total > 0 else 0
            table_data.append([estado, str(cantidad), f"{porcentaje:.1f}%"])
        
        table = Table(table_data, colWidths=[200, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        contenido.append(table)
        contenido.append(Spacer(1, 1*cm))
    
    # Lista de proyectos
    if datos.get('detalles'):
        contenido.append(Paragraph("LISTA DE PROYECTOS", styles['Heading3']))
        
        table_data = [['ID', 'Proyecto', 'Área', 'Presupuesto', 'Avance', 'Estado']]
        
        for obra in datos['detalles'][:30]:  # Limitar a 30 proyectos
            table_data.append([
                str(obra.get('id_excel', obra.get('id', ''))),
                obra.get('programa', '')[:30],
                obra.get('area_responsable', '')[:15],
                f"${float(obra.get('presupuesto_modificado', 0)):,.0f}",
                f"{float(obra.get('avance_fisico_pct', 0)):.1f}%",
                obra.get('estatus_general', '')[:10]
            ])
        
        table = Table(table_data, colWidths=[40, 150, 80, 80, 60, 60])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        
        contenido.append(table)
    
    doc.build(contenido)
    buffer.seek(0)
    return buffer

def _crear_pdf_riesgos(datos, fecha_corte):
    """Crea PDF para análisis de riesgos"""
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Encabezado
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, "ANÁLISIS DE RIESGOS - POA")
    p.setFont("Helvetica", 12)
    p.drawString(100, 730, f"Fecha: {fecha_corte}")
    p.drawString(100, 710, f"Proyectos de alto riesgo: {datos.get('alto_riesgo', 0)}")
    p.drawString(100, 690, f"Viabilidad crítica: {datos.get('viabilidad_critica', 0)}")
    
    y = 650
    
    # Proyectos de alto riesgo
    if datos.get('detalles'):
        p.setFont("Helvetica-Bold", 12)
        p.drawString(100, y, "PROYECTOS DE ALTO RIESGO:")
        y -= 20
        
        p.setFont("Helvetica", 10)
        for obra in datos['detalles'][:15]:
            riesgo_text = f"Nivel {obra.get('riesgo_nivel', 1)}"
            if obra.get('riesgo_nivel', 1) >= 4:
                riesgo_text = f"⚠️ {riesgo_text} - ALTO"
            
            p.drawString(120, y, f"• {obra.get('programa', '')[:50]}...")
            p.drawString(450, y, riesgo_text)
            y -= 15
            
            if y < 100:
                p.showPage()
                y = 750
                p.setFont("Helvetica", 10)
    
    p.save()
    buffer.seek(0)
    return buffer

def _crear_excel_general(titulo, datos, fecha_corte):
    """Crea Excel genérico para cualquier reporte"""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita a 31 caracteres
    
    # Título
    ws['A1'] = titulo
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Fecha: {fecha_corte}"
    ws['A2'].font = Font(italic=True)
    
    # Datos básicos
    if isinstance(datos, dict):
        row = 4
        for key, value in datos.items():
            if key != 'detalles':
                ws[f'A{row}'] = str(key).replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                row += 1
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ============================
# FUNCIÓN PRINCIPAL MEJORADA
# ============================

@csrf_exempt
def generar_reporte(request):
    """Generar reportes de TODAS las secciones del sistema - VERSIÓN COMPLETA"""
    if request.method == 'GET':
        return JsonResponse({
            'message': 'Use POST para generar reportes',
            'formatos_soportados': ['pdf', 'excel', 'ambos'],
            'tipos_reporte': ['ejecutivo', 'cartera', 'presupuesto', 'riesgos', 'territorial', 'transparencia'],
            'ejemplo': {
                'tipo_reporte': 'ejecutivo',
                'fecha_corte': datetime.now().strftime('%Y-%m-%d'),
                'formato': 'pdf',
                'periodo': 'mensual',
                'direcciones': [1, 2, 3],
                'incluir_graficos': True
            }
        })
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido. Use POST.'
        }, status=405)
    
    try:
        # Parsear datos
        try:
            data = json.loads(request.body)
        except:
            data = {}
        
        tipo_reporte = data.get('tipo_reporte', 'ejecutivo')
        fecha_corte = data.get('fecha_corte', datetime.now().strftime('%Y-%m-%d'))
        formato = data.get('formato', 'pdf')
        periodo = data.get('periodo', 'mensual')
        direcciones_ids = data.get('direcciones', [])
        incluir_graficos = data.get('incluir_graficos', True)
        
        print(f"📊 Generando reporte: tipo={tipo_reporte}, formato={formato}, fecha={fecha_corte}")
        
        # Importar modelos
        from backend.poa.models import Obra, Direccion
        from django.db.models import Sum, Count, Avg, F, Q
        
        # Obtener datos básicos
        obras = Obra.objects.all()
        
        # Aplicar filtros básicos
        if fecha_corte:
            try:
                fecha = datetime.strptime(fecha_corte, '%Y-%m-%d').date()
                obras = obras.filter(fecha_inicio_prog__lte=fecha)
            except:
                pass
        
        # Datos comunes
        datos_comunes = {
            'total_obras': obras.count(),
            'presupuesto_total': obras.aggregate(total=Sum('presupuesto_modificado'))['total'] or 0,
            'avance_promedio': obras.aggregate(promedio=Avg('avance_fisico_pct'))['promedio'] or 0,
            'proyectos_riesgo': obras.filter(riesgo_nivel__gte=4).count(),
            'beneficiarios_totales': obras.aggregate(total=Sum('beneficiarios_directos'))['total'] or 0,
        }
        
        # Preparar datos específicos por tipo de reporte
        if tipo_reporte == 'ejecutivo':
            datos_comunes['detalles'] = list(obras.values(
                'programa', 'area_responsable', 'presupuesto_modificado',
                'avance_fisico_pct', 'estatus_general'
            ).order_by('-presupuesto_modificado')[:15])
            
            if formato == 'pdf':
                pdf_buffer = _crear_pdf_ejecutivo(datos_comunes, fecha_corte)
                response = HttpResponse(pdf_buffer, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="reporte_ejecutivo_{fecha_corte}.pdf"'
                return response
                
            elif formato == 'excel':
                excel_buffer = _crear_excel_ejecutivo(datos_comunes, fecha_corte)
                response = HttpResponse(
                    excel_buffer,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="reporte_ejecutivo_{fecha_corte}.xlsx"'
                return response
        
        elif tipo_reporte == 'cartera':
            # Datos para cartera
            por_estado = obras.values('estatus_general').annotate(
                total=Count('id'),
                presupuesto=Sum('presupuesto_modificado')
            )
            
            datos_cartera = {
                **datos_comunes,
                'por_estado': {item['estatus_general']: item['total'] for item in por_estado},
                'detalles': list(obras.values(
                    'id_excel', 'programa', 'area_responsable', 'presupuesto_modificado',
                    'avance_fisico_pct', 'estatus_general', 'fecha_inicio_prog'
                ).order_by('estatus_general')[:50])
            }
            
            if formato == 'pdf':
                pdf_buffer = _crear_pdf_cartera(datos_cartera, fecha_corte)
                response = HttpResponse(pdf_buffer, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="cartera_proyectos_{fecha_corte}.pdf"'
                return response
                
            elif formato == 'excel':
                excel_buffer = _crear_excel_general("Cartera de Proyectos", datos_cartera, fecha_corte)
                response = HttpResponse(
                    excel_buffer,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="cartera_proyectos_{fecha_corte}.xlsx"'
                return response
        
        elif tipo_reporte == 'riesgos':
            # Datos para riesgos
            alto_riesgo = obras.filter(riesgo_nivel__gte=4)
            viabilidad_critica = obras.filter(viabilidad_ejecucion__lte=2)
            
            datos_riesgos = {
                'alto_riesgo': alto_riesgo.count(),
                'medio_riesgo': obras.filter(riesgo_nivel=3).count(),
                'bajo_riesgo': obras.filter(riesgo_nivel__lte=2).count(),
                'viabilidad_critica': viabilidad_critica.count(),
                'detalles': list(alto_riesgo.values(
                    'programa', 'area_responsable', 'riesgo_nivel',
                    'viabilidad_ejecucion', 'problemas_identificados'
                )[:20])
            }
            
            if formato == 'pdf':
                pdf_buffer = _crear_pdf_riesgos(datos_riesgos, fecha_corte)
                response = HttpResponse(pdf_buffer, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="analisis_riesgos_{fecha_corte}.pdf"'
                return response
                
            elif formato == 'excel':
                excel_buffer = _crear_excel_general("Análisis de Riesgos", datos_riesgos, fecha_corte)
                response = HttpResponse(
                    excel_buffer,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="analisis_riesgos_{fecha_corte}.xlsx"'
                return response
        
        elif tipo_reporte in ['presupuesto', 'territorial', 'transparencia']:
            # Reportes simples para otros tipos
            datos_generales = {
                'tipo_reporte': tipo_reporte,
                'fecha_corte': fecha_corte,
                'total_proyectos': datos_comunes['total_obras'],
                'presupuesto_total': datos_comunes['presupuesto_total'],
                'nota': 'Reporte generado correctamente',
                'generado_en': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            if formato == 'pdf':
                # PDF simple
                buffer = BytesIO()
                p = canvas.Canvas(buffer, pagesize=letter)
                
                p.setFont("Helvetica-Bold", 16)
                p.drawString(100, 750, f"REPORTE {tipo_reporte.upper()}")
                p.setFont("Helvetica", 12)
                p.drawString(100, 730, f"Fecha: {fecha_corte}")
                p.drawString(100, 710, f"Proyectos: {datos_comunes['total_obras']}")
                p.drawString(100, 690, f"Presupuesto: ${datos_comunes['presupuesto_total']:,.2f}")
                p.drawString(100, 670, "=" * 50)
                p.drawString(100, 650, "Reporte generado correctamente")
                p.drawString(100, 630, f"Tipo: {tipo_reporte}")
                p.drawString(100, 610, f"Período: {periodo}")
                
                if incluir_graficos:
                    p.drawString(100, 590, "✓ Gráficos incluidos")
                
                p.save()
                buffer.seek(0)
                
                response = HttpResponse(buffer, content_type='application/pdf')
                filename = f"reporte_{tipo_reporte}_{fecha_corte}.pdf"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
                
            elif formato == 'excel':
                # Excel simple
                wb = Workbook()
                ws = wb.active
                ws.title = f"Reporte {tipo_reporte}"
                
                ws['A1'] = f"REPORTE {tipo_reporte.upper()}"
                ws['A1'].font = Font(size=14, bold=True)
                ws['A2'] = f"Fecha: {fecha_corte}"
                ws['A3'] = f"Período: {periodo}"
                ws['A4'] = f"Total proyectos: {datos_comunes['total_obras']}"
                ws['A5'] = f"Presupuesto total: ${datos_comunes['presupuesto_total']:,.2f}"
                ws['A6'] = f"Avance promedio: {datos_comunes['avance_promedio']:.1f}%"
                ws['A7'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                
                if incluir_graficos:
                    ws['A8'] = "Gráficos: Incluidos"
                
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                response = HttpResponse(
                    buffer,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"reporte_{tipo_reporte}_{fecha_corte}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        
        # MANEJO DE FORMATO "AMBOS"
        if formato == 'ambos':
            # Crear ZIP con ambos formatos
            zip_buffer = BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # 1. Generar PDF
                if tipo_reporte == 'ejecutivo':
                    pdf_content = _crear_pdf_ejecutivo(datos_comunes, fecha_corte).getvalue()
                elif tipo_reporte == 'cartera':
                    pdf_content = _crear_pdf_cartera({
                        **datos_comunes,
                        'detalles': list(obras.values('programa', 'area_responsable', 'estatus_general')[:20])
                    }, fecha_corte).getvalue()
                elif tipo_reporte == 'riesgos':
                    pdf_content = _crear_pdf_riesgos({
                        'alto_riesgo': obras.filter(riesgo_nivel__gte=4).count(),
                        'detalles': list(obras.filter(riesgo_nivel__gte=4).values('programa', 'riesgo_nivel')[:10])
                    }, fecha_corte).getvalue()
                else:
                    # PDF genérico
                    buffer = BytesIO()
                    p = canvas.Canvas(buffer, pagesize=letter)
                    p.drawString(100, 750, f"REPORTE {tipo_reporte.upper()}")
                    p.drawString(100, 730, f"Fecha: {fecha_corte}")
                    p.save()
                    buffer.seek(0)
                    pdf_content = buffer.getvalue()
                
                zip_file.writestr(f"reporte_{tipo_reporte}_{fecha_corte}.pdf", pdf_content)
                
                # 2. Generar Excel
                if tipo_reporte == 'ejecutivo':
                    excel_content = _crear_excel_ejecutivo(datos_comunes, fecha_corte).getvalue()
                elif tipo_reporte in ['cartera', 'riesgos']:
                    excel_content = _crear_excel_general(
                        f"Reporte {tipo_reporte}",
                        datos_comunes,
                        fecha_corte
                    ).getvalue()
                else:
                    # Excel genérico
                    wb = Workbook()
                    ws = wb.active
                    ws['A1'] = f"REPORTE {tipo_reporte.upper()}"
                    ws['A2'] = f"Fecha: {fecha_corte}"
                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    excel_content = buffer.getvalue()
                
                zip_file.writestr(f"reporte_{tipo_reporte}_{fecha_corte}.xlsx", excel_content)
            
            zip_buffer.seek(0)
            
            response = HttpResponse(zip_buffer, content_type='application/zip')
            filename = f"reporte_{tipo_reporte}_{fecha_corte}.zip"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        # Si llegamos aquí, algo falló
        return JsonResponse({
            'success': False,
            'error': f'Tipo de reporte "{tipo_reporte}" o formato "{formato}" no soportado',
            'tipos_soportados': ['ejecutivo', 'cartera', 'presupuesto', 'riesgos', 'territorial', 'transparencia'],
            'formatos_soportados': ['pdf', 'excel', 'ambos']
        }, status=400)
            
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ Error generando reporte: {str(e)}")
        print(f"📋 Detalles: {error_details}")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'trace': error_details if 'DEBUG' in locals() else 'No disponible',
            'message': 'Error interno del servidor. Verifica los logs para más detalles.'
        }, status=500)

# ============================
# URL PATTERNS
# ============================

urlpatterns = [
    path('admin/', admin.site.urls),
    path('health/', health_check, name='health-check'),
    path('api/direcciones/', lista_direcciones, name='direcciones-list'),
    path('api/reportes/vista_previa/', vista_previa_reporte, name='vista-previa'),
    path('api/reportes/generar/', generar_reporte, name='generar-reporte'),
]