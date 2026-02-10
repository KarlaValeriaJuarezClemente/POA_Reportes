# Proyecto\POA_Reporte\backend\poa\generador.py

import os
import tempfile
from datetime import datetime
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ConfiguracionReporte:
    """Clase simple para configuración de reportes"""
    def __init__(self, **kwargs):
        self.nombre = kwargs.get('nombre', 'Reporte POA')
        self.tipo_reporte = kwargs.get('tipo_reporte', 'ejecutivo')
        self.periodo = kwargs.get('periodo', 'mensual')
        self.fecha_corte = kwargs.get('fecha_corte', datetime.now().date())
        self.formato_salida = kwargs.get('formato_salida', 'pdf')
        self.incluir_graficos = kwargs.get('incluir_graficos', True)
        self.incluir_anexos = kwargs.get('incluir_anexos', False)


class GeneradorReportes:
    """Clase para generar reportes en diferentes formatos"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._configurar_estilos()
    
    def _configurar_estilos(self):
        """Configura estilos personalizados para los reportes"""
        # Título principal
        self.styles.add(ParagraphStyle(
            name='TituloPrincipal',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=20,
            textColor=colors.HexColor('#1e3a8a'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Subtítulo
        self.styles.add(ParagraphStyle(
            name='Subtitulo',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            textColor=colors.HexColor('#374151'),
            alignment=TA_CENTER
        ))
        
        # Encabezado de sección
        self.styles.add(ParagraphStyle(
            name='Seccion',
            parent=self.styles['Heading3'],
            fontSize=12,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor('#1e40af'),
            fontName='Helvetica-Bold'
        ))
        
        # Texto normal
        self.styles.add(ParagraphStyle(
            name='TextoNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=8,
            textColor=colors.black
        ))
        
        # Texto resaltado
        self.styles.add(ParagraphStyle(
            name='Resaltado',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#059669'),
            backColor=colors.HexColor('#f0fdf4'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        ))
    
    def generar_pdf_reporte(self, obras, config, estadisticas):
        """Genera reporte en formato PDF"""
        try:
            # Crear archivo temporal
            temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            
            # Configurar documento según tipo de reporte
            if config.tipo_reporte in ['cartera', 'presupuesto']:
                doc = SimpleDocTemplate(
                    temp_file.name,
                    pagesize=landscape(A4),
                    topMargin=1*cm,
                    bottomMargin=1*cm,
                    leftMargin=1.5*cm,
                    rightMargin=1.5*cm
                )
            else:
                doc = SimpleDocTemplate(
                    temp_file.name,
                    pagesize=A4,
                    topMargin=2*cm,
                    bottomMargin=2*cm,
                    leftMargin=2*cm,
                    rightMargin=2*cm
                )
            
            elements = []
            
            # 1. Encabezado
            elements.extend(self._crear_encabezado(config))
            
            # 2. Resumen ejecutivo
            if config.tipo_reporte == 'ejecutivo':
                elements.extend(self._crear_resumen_ejecutivo(estadisticas))
            
            # 3. Tabla de datos
            elements.extend(self._crear_tabla_datos(obras, config))
            
            # 4. Análisis y gráficos
            if config.incluir_graficos:
                elements.extend(self._crear_analisis(estadisticas, config))
            
            # 5. Pie de página
            elements.extend(self._crear_pie_pagina())
            
            # Generar PDF
            doc.build(elements)
            
            return temp_file.name
            
        except Exception as e:
            print(f"Error al generar PDF: {str(e)}")
            raise
    
    def _crear_encabezado(self, config):
        """Crea el encabezado del reporte"""
        elementos = []
        
        # Logo (si existe)
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
        if os.path.exists(logo_path):
            try:
                elementos.append(Image(logo_path, width=3*cm, height=1.5*cm))
                elementos.append(Spacer(1, 0.5*cm))
            except:
                pass
        
        # Título
        titulos = {
            'ejecutivo': 'REPORTE EJECUTIVO - POA 2026',
            'cartera': 'CARTERA DE PROYECTOS',
            'presupuesto': 'EJECUCIÓN PRESUPUESTAL',
            'riesgos': 'ANÁLISIS DE RIESGOS',
            'territorial': 'IMPACTO TERRITORIAL',
        }
        
        titulo = titulos.get(config.tipo_reporte, 'REPORTE POA 2026')
        elementos.append(Paragraph(titulo, self.styles['TituloPrincipal']))
        
        # Subtítulo
        try:
            periodo_display = self._get_periodo_display(config.periodo)
            if hasattr(config.fecha_corte, 'strftime'):
                fecha_str = config.fecha_corte.strftime('%d de %B de %Y')
            else:
                fecha_str = str(config.fecha_corte)
            subtitulo = f"{periodo_display.upper()} - CORTE AL {fecha_str}"
        except:
            if hasattr(config.fecha_corte, 'strftime'):
                subtitulo = f"CORTE AL {config.fecha_corte.strftime('%d/%m/%Y')}"
            else:
                subtitulo = f"CORTE AL {config.fecha_corte}"
            
        elementos.append(Paragraph(subtitulo, self.styles['Subtitulo']))
        
        # Información del reporte
        try:
            if hasattr(config.fecha_corte, 'strftime'):
                fecha_display = config.fecha_corte.strftime('%d/%m/%Y')
            else:
                fecha_display = str(config.fecha_corte)
                
            info_texto = f"""
            <b>Tipo:</b> {self._get_tipo_reporte_display(config.tipo_reporte)} | 
            <b>Período:</b> {self._get_periodo_display(config.periodo)} | 
            <b>Fecha de corte:</b> {fecha_display}
            """
            elementos.append(Paragraph(info_texto, self.styles['TextoNormal']))
        except:
            pass
        
        elementos.append(Spacer(1, 1*cm))
        return elementos
    
    def _crear_resumen_ejecutivo(self, estadisticas):
        """Crea sección de resumen ejecutivo"""
        elementos = []
        
        elementos.append(Paragraph("RESUMEN EJECUTIVO", self.styles['Seccion']))
        
        # Formatear números con separadores de miles
        total_obras = estadisticas.get('total_obras', 0)
        presupuesto_total = estadisticas.get('presupuesto_total', 0)
        beneficiarios_total = estadisticas.get('beneficiarios_total', 0)
        avance_promedio = estadisticas.get('avance_promedio', 0)
        presupuesto_promedio = estadisticas.get('presupuesto_promedio', 0)
        
        resumen_texto = f"""
        <b>Total de Proyectos:</b> {total_obras:,}<br/>
        <b>Presupuesto Total:</b> ${presupuesto_total:,.2f}<br/>
        <b>Beneficiarios Directos:</b> {beneficiarios_total:,}<br/>
        <b>Avance Promedio:</b> {avance_promedio:.1f}%<br/>
        <b>Presupuesto Promedio por Proyecto:</b> ${presupuesto_promedio:.2f}<br/>
        """
        
        elementos.append(Paragraph(resumen_texto, self.styles['Resaltado']))
        elementos.append(Spacer(1, 0.5*cm))
        
        return elementos
    
    def _crear_tabla_datos(self, obras, config):
        """Crea tabla de datos según tipo de reporte"""
        elementos = []
        
        elementos.append(Paragraph("DETALLE DE PROYECTOS", self.styles['Seccion']))
        
        # Preparar datos según tipo de reporte
        data = []
        
        if config.tipo_reporte == 'ejecutivo':
            headers = ['Proyecto', 'Área', 'Presupuesto', 'Avance', 'Estado']
            data.append(headers)
            
            for obra in obras[:50]:  # Limitar a 50 para evitar PDFs muy grandes
                presupuesto = getattr(obra, 'presupuesto_modificado', 0) or 0
                if presupuesto <= 0:
                    presupuesto = getattr(obra, 'anteproyecto_total', 0) or 0
                
                programa = getattr(obra, 'programa', 'Sin nombre') or 'Sin nombre'
                area = getattr(obra, 'area_responsable', 'N/A') or 'N/A'
                avance = getattr(obra, 'avance_fisico_pct', 0) or 0
                estado = getattr(obra, 'estatus_general', 'N/A') or 'N/A'
                
                data.append([
                    str(programa)[:50] + '...' if len(str(programa)) > 50 else str(programa),
                    str(area)[:30],
                    f"${float(presupuesto):,.2f}" if presupuesto else "N/A",
                    f"{float(avance):.1f}%",
                    str(estado)[:20]
                ])
        
        elif config.tipo_reporte == 'cartera':
            headers = ['ID', 'Proyecto', 'Área', 'Presupuesto', 'Avance']
            data.append(headers)
            
            for obra in obras[:50]:
                presupuesto = getattr(obra, 'presupuesto_modificado', 0) or 0
                if presupuesto <= 0:
                    presupuesto = getattr(obra, 'anteproyecto_total', 0) or 0
                
                id_excel = getattr(obra, 'id_excel', obra.id)
                programa = getattr(obra, 'programa', 'Sin nombre') or 'Sin nombre'
                area = getattr(obra, 'area_responsable', 'N/A') or 'N/A'
                avance = getattr(obra, 'avance_fisico_pct', 0) or 0
                
                data.append([
                    str(id_excel),
                    str(programa)[:40] + '...' if len(str(programa)) > 40 else str(programa),
                    str(area)[:20],
                    f"${float(presupuesto):,.2f}" if presupuesto else "N/A",
                    f"{float(avance):.1f}%"
                ])
        
        elif config.tipo_reporte == 'presupuesto':
            headers = ['Proyecto', 'Área', 'Presupuesto', 'Ejecutado', '% Ejecución']
            data.append(headers)
            
            for obra in obras[:50]:
                presupuesto_mod = getattr(obra, 'presupuesto_modificado', 0) or 0
                anteproyecto = getattr(obra, 'anteproyecto_total', 0) or 0
                avance_financiero = getattr(obra, 'avance_financiero_pct', 0) or 0
                
                presupuesto = presupuesto_mod if presupuesto_mod > 0 else anteproyecto
                ejecutado = float(presupuesto) * (float(avance_financiero) / 100) if presupuesto and avance_financiero else 0
                
                programa = getattr(obra, 'programa', 'Sin nombre') or 'Sin nombre'
                area = getattr(obra, 'area_responsable', 'N/A') or 'N/A'
                
                data.append([
                    str(programa)[:40] + '...' if len(str(programa)) > 40 else str(programa),
                    str(area)[:20],
                    f"${float(presupuesto):,.2f}" if presupuesto else "N/A",
                    f"${ejecutado:,.2f}",
                    f"{float(avance_financiero):.1f}%"
                ])
        
        elif config.tipo_reporte == 'riesgos':
            headers = ['Proyecto', 'Área', 'Riesgo', 'Viabilidad']
            data.append(headers)
            
            for obra in obras[:50]:
                programa = getattr(obra, 'programa', 'Sin nombre') or 'Sin nombre'
                area = getattr(obra, 'area_responsable', 'N/A') or 'N/A'
                riesgo = getattr(obra, 'riesgo_nivel', 1) or 1
                viabilidad = getattr(obra, 'viabilidad_ejecucion', 1) or 1
                
                data.append([
                    str(programa)[:40] + '...' if len(str(programa)) > 40 else str(programa),
                    str(area)[:20],
                    self._get_riesgo_text(riesgo),
                    self._get_viabilidad_text(viabilidad)
                ])
        
        # Crear tabla solo si hay datos
        if len(data) > 1:
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            elementos.append(table)
        
        elementos.append(Spacer(1, 1*cm))
        return elementos
    
    def _crear_analisis(self, estadisticas, config):
        """Crea sección de análisis"""
        elementos = []
        
        if config.tipo_reporte in ['ejecutivo', 'presupuesto', 'riesgos']:
            elementos.append(Paragraph("ANÁLISIS ESTADÍSTICO", self.styles['Seccion']))
            
            # Distribución por estado
            if estadisticas.get('obras_por_estado'):
                elementos.append(Paragraph("<b>Distribución por Estado:</b>", self.styles['TextoNormal']))
                
                estados_texto = ""
                total_obras = estadisticas.get('total_obras', 1)
                for estado, cantidad in list(estadisticas['obras_por_estado'].items())[:10]:
                    porcentaje = (cantidad / total_obras * 100) if total_obras > 0 else 0
                    estados_texto += f"{estado}: {cantidad} ({porcentaje:.1f}%)<br/>"
                
                elementos.append(Paragraph(estados_texto, self.styles['TextoNormal']))
        
        elementos.append(Spacer(1, 0.5*cm))
        return elementos
    
    def _crear_pie_pagina(self):
        """Crea pie de página"""
        elementos = []
        
        elementos.append(Spacer(1, 1*cm))
        elementos.append(Paragraph(
            f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | Sistema POA 2026",
            ParagraphStyle(
                name='PiePagina',
                parent=self.styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER
            )
        ))
        
        return elementos
    
    def generar_excel_reporte(self, obras, config, estadisticas):
        """Genera reporte en formato Excel"""
        try:
            # Crear archivo temporal
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            
            wb = Workbook()
            
            # Hoja de resumen
            ws_resumen = wb.active
            ws_resumen.title = "Resumen"
            
            # Estilos
            header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezado
            ws_resumen.merge_cells('A1:E1')
            ws_resumen['A1'] = f"REPORTE {self._get_tipo_reporte_display(config.tipo_reporte).upper()} - POA 2026"
            ws_resumen['A1'].font = Font(size=16, bold=True)
            ws_resumen['A1'].alignment = center_alignment
            
            fecha_display = config.fecha_corte.strftime('%d/%m/%Y') if hasattr(config.fecha_corte, 'strftime') else str(config.fecha_corte)
            
            ws_resumen['A2'] = f"Período: {self._get_periodo_display(config.periodo)}"
            ws_resumen['B2'] = f"Fecha de corte: {fecha_display}"
            ws_resumen['C2'] = f"Total proyectos: {estadisticas.get('total_obras', 0)}"
            ws_resumen['D2'] = f"Presupuesto total: ${estadisticas.get('presupuesto_total', 0):,.2f}"
            ws_resumen['E2'] = f"Beneficiarios: {estadisticas.get('beneficiarios_total', 0):,}"
            
            # Estadísticas
            ws_resumen['A4'] = "ESTADÍSTICAS DEL REPORTE"
            ws_resumen['A4'].font = Font(bold=True, size=12)
            
            row = 5
            stats = [
                ['Total de Proyectos', estadisticas.get('total_obras', 0)],
                ['Presupuesto Total', f"${estadisticas.get('presupuesto_total', 0):,.2f}"],
                ['Beneficiarios Totales', estadisticas.get('beneficiarios_total', 0)],
                ['Avance Promedio', f"{estadisticas.get('avance_promedio', 0):.1f}%"],
                ['Presupuesto Promedio', f"${estadisticas.get('presupuesto_promedio', 0):,.2f}"]
            ]
            
            for stat in stats:
                ws_resumen[f'A{row}'] = stat[0]
                ws_resumen[f'B{row}'] = stat[1]
                row += 1
            
            # Hoja de datos detallados
            ws_detalle = wb.create_sheet("Detalle")
            
            # Encabezados según tipo de reporte
            if config.tipo_reporte == 'ejecutivo':
                headers = ['ID', 'Proyecto', 'Área Responsable', 'Presupuesto', 'Avance Físico', 'Estado']
            elif config.tipo_reporte == 'presupuesto':
                headers = ['ID', 'Proyecto', 'Área', 'Presupuesto Modificado', 'Anteproyecto', 'Avance Financiero %']
            elif config.tipo_reporte == 'riesgos':
                headers = ['ID', 'Proyecto', 'Área', 'Riesgo Nivel', 'Viabilidad Ejecución']
            else:  # cartera o por defecto
                headers = ['ID', 'Proyecto', 'Área', 'Presupuesto', 'Avance Físico', 'Estado']
            
            # Escribir encabezados
            for col, header in enumerate(headers, 1):
                cell = ws_detalle.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Escribir datos
            row = 2
            for obra in obras:
                for col, header in enumerate(headers, 1):
                    value = self._get_excel_cell_value(obra, header, config.tipo_reporte)
                    cell = ws_detalle.cell(row=row, column=col, value=value)
                    cell.border = border
                
                row += 1
            
            # Ajustar anchos de columna
            for column in ws_detalle.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_detalle.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(temp_file.name)
            
            return temp_file.name
            
        except Exception as e:
            print(f"Error al generar Excel: {str(e)}")
            raise
    
    def _get_excel_cell_value(self, obra, header, tipo_reporte):
        """Obtiene el valor para una celda de Excel según el header"""
        try:
            if header == 'ID':
                return getattr(obra, 'id_excel', obra.id)
            elif header == 'Proyecto':
                return str(getattr(obra, 'programa', 'Sin nombre') or 'Sin nombre')
            elif header in ['Área', 'Área Responsable']:
                return str(getattr(obra, 'area_responsable', 'N/A') or 'N/A')
            elif header == 'Presupuesto':
                presupuesto = getattr(obra, 'presupuesto_modificado', 0) or 0
                if presupuesto <= 0:
                    presupuesto = getattr(obra, 'anteproyecto_total', 0) or 0
                return float(presupuesto)
            elif header == 'Avance Físico':
                return float(getattr(obra, 'avance_fisico_pct', 0) or 0)
            elif header == 'Estado':
                return str(getattr(obra, 'estatus_general', 'N/A') or 'N/A')
            elif header == 'Riesgo Nivel':
                return int(getattr(obra, 'riesgo_nivel', 1) or 1)
            elif header == 'Viabilidad Ejecución':
                return int(getattr(obra, 'viabilidad_ejecucion', 1) or 1)
            elif header == 'Presupuesto Modificado':
                return float(getattr(obra, 'presupuesto_modificado', 0) or 0)
            elif header == 'Anteproyecto':
                return float(getattr(obra, 'anteproyecto_total', 0) or 0)
            elif header == 'Avance Financiero %':
                return float(getattr(obra, 'avance_financiero_pct', 0) or 0)
            else:
                return ''
        except Exception as e:
            print(f"Error obteniendo valor para {header}: {str(e)}")
            return ''
    
    def _get_riesgo_text(self, nivel):
        """Convierte nivel de riesgo numérico a texto"""
        niveles = {
            1: "Muy Bajo",
            2: "Bajo", 
            3: "Medio",
            4: "Alto",
            5: "Muy Alto"
        }
        try:
            nivel_int = int(nivel)
            return niveles.get(nivel_int, f"Nivel {nivel_int}")
        except:
            return "N/A"
    
    def _get_viabilidad_text(self, nivel):
        """Convierte viabilidad numérica a texto"""
        niveles = {
            1: "Muy Baja",
            2: "Baja",
            3: "Media",
            4: "Alta",
            5: "Muy Alta"
        }
        try:
            nivel_int = int(nivel)
            return niveles.get(nivel_int, f"Nivel {nivel_int}")
        except:
            return "N/A"
    
    def _get_tipo_reporte_display(self, tipo_reporte):
        """Obtiene el nombre display del tipo de reporte"""
        tipos = {
            'ejecutivo': 'Ejecutivo',
            'cartera': 'Cartera de Proyectos',
            'presupuesto': 'Ejecución Presupuestal',
            'riesgos': 'Análisis de Riesgos',
            'territorial': 'Impacto Territorial',
        }
        return tipos.get(tipo_reporte, tipo_reporte)
    
    def _get_periodo_display(self, periodo):
        """Obtiene el nombre display del período"""
        periodos = {
            'semanal': 'Semanal',
            'mensual': 'Mensual',
            'trimestral': 'Trimestral',
            'anual': 'Anual',
        }
        return periodos.get(periodo, periodo)